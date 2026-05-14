from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ── 公共样式常量 ──
HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
HEADER_FONT = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=11)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGNMENT = Alignment(horizontal="center", vertical="center")
SUCCESS_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
FAILED_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
SUCCESS_FONT = Font(color="059669", bold=True)
FAILED_FONT = Font(color="DC2626", bold=True)


def _style_header(ws, col_widths: dict[int, int], header_row: int = 1) -> None:
    """设置指定行的表头样式、列宽、冻结首行和自动筛选。"""
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for cell in ws[header_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = ws.dimensions


def build_attendance_workbook(records) -> BytesIO:
    """把考勤记录导出为格式化的 Excel 二进制流。"""

    wb = Workbook()
    ws = wb.active
    ws.title = "考勤记录"
    headers = ["时间", "课程", "学号", "姓名", "班级", "状态", "置信度", "活体分", "情绪", "备注"]
    ws.append(headers)

    for item in records:
        student = item.student
        status_text = "成功" if item.status == "success" else "失败"
        ws.append(
            [
                item.timestamp.strftime("%Y-%m-%d %H:%M:%S"),
                item.course_name,
                student.student_no if student else "",
                student.name if student else "",
                student.class_name if student else "",
                status_text,
                round(item.confidence, 3) if item.confidence else 0,
                item.liveness_score,
                item.emotion_type or "",
                item.message or "",
            ]
        )

    _style_header(ws, {1: 20, 2: 14, 3: 16, 4: 10, 5: 16, 6: 8, 7: 10, 8: 10, 9: 10, 10: 30})

    # 状态列着色（第 6 列）
    for row in ws.iter_rows(min_row=2, max_col=len(headers), values_only=False):
        status_cell = row[5]  # 第 6 列（0-indexed: 5）
        if status_cell.value == "成功":
            status_cell.fill = SUCCESS_FILL
            status_cell.font = SUCCESS_FONT
        elif status_cell.value == "失败":
            status_cell.fill = FAILED_FILL
            status_cell.font = FAILED_FONT
        status_cell.alignment = CELL_ALIGNMENT

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def build_attendance_stats_workbook(students: list[dict], dates: list[str]) -> BytesIO:
    """把课程考勤统计导出为格式化的 Excel。

    students: [{name, class_name, checkin: {date: bool}}, ...]
    dates:    ["2026-05-11", "2026-05-12", ...]
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "考勤统计"

    # 汇总行
    total = len(students)
    signed_total = sum(1 for s in students if any(s.get("checkin", {}).get(d) for d in dates))
    rate = f"{signed_total / total * 100:.1f}%" if total else "0%"
    ws.append(["到课率", rate, f"应到 {total} 人", f"实到 {signed_total} 人", f"缺勤 {total - signed_total} 人"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    # 汇总行样式
    for cell in ws[1]:
        cell.font = Font(bold=True, size=11)

    if len(dates) == 1:
        absent_names = [s["name"] for s in students if not s.get("checkin", {}).get(dates[0])]
        if absent_names:
            ws.append(["未签到", "、".join(absent_names)])
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)

    ws.append([])  # 空行分隔

    # 表头
    row_idx = ws.max_row + 1
    headers = ["姓名", "班级"]
    for d in dates:
        headers.append(d)
    headers.append("总签到次数")
    ws.append(headers)
    _style_header(ws, {1: 10, 2: 14} | {i: 12 for i in range(3, len(headers) + 1)}, header_row=row_idx)

    # 数据行
    for student in students:
        row = [student["name"], student.get("class_name", "")]
        total_count = 0
        for d in dates:
            checked = student.get("checkin", {}).get(d, False)
            row.append("✓" if checked else "")
            if checked:
                total_count += 1
        row.append(total_count)
        ws.append(row)

    # 数据行居中对齐
    for row in ws.iter_rows(min_row=row_idx + 1, max_col=len(headers), values_only=False):
        for cell in row:
            cell.alignment = CELL_ALIGNMENT

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def build_activity_workbook(records: list[dict]) -> BytesIO:
    """把活动参与记录导出为 Excel。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "活动参与名单"

    headers = ["活动名称", "日期", "学号", "姓名", "班级", "置信度", "情绪"]
    ws.append(headers)

    for r in records:
        ws.append([
            r["activity_name"],
            r["activity_date"],
            r["student_no"],
            r["name"],
            r["class_name"],
            r["confidence"],
            r.get("emotion") or "",
        ])

    _style_header(ws, {1: 20, 2: 14, 3: 16, 4: 10, 5: 16, 6: 12, 7: 10})

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
